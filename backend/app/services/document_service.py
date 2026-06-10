"""Office document conversions.

Office->PDF conversions (Word/Excel/PowerPoint -> PDF) use LibreOffice (the
`soffice` binary) in headless mode. PDF->Word uses pdf2docx and PDF->Excel
uses pdfplumber + openpyxl — LibreOffice imports PDFs as *Draw* documents that
cannot be re-exported through the Writer/Calc filters, so dedicated PDF
parsers are required for the PDF->Office direction.
"""
from __future__ import annotations

from pathlib import Path

import pdfplumber
from openpyxl import Workbook

from app.config import settings
from app.core.errors import ProcessingError
from app.schemas.jobs import JobResult
from app.services.base import ensure_binary, file_result, run_command, stem


def _libreoffice_convert(src: Path, out_dir: Path, target_format: str, target_ext: str) -> Path:
    """Convert `src` to `target_format` using a headless LibreOffice instance."""
    soffice = ensure_binary(settings.libreoffice_cmd)
    # Isolated user profile so concurrent conversions don't clash.
    profile = out_dir / ".loprofile"
    profile.mkdir(parents=True, exist_ok=True)
    run_command(
        [
            soffice,
            f"-env:UserInstallation={profile.as_uri()}",
            "--headless",
            "--norestore",
            "--nolockcheck",
            "--convert-to",
            target_format,
            "--outdir",
            str(out_dir),
            str(src),
        ],
        timeout=300,
    )
    produced = out_dir / f"{stem(src.name)}{target_ext}"
    if produced.is_file():
        return produced
    candidates = sorted(out_dir.glob(f"*{target_ext}"))
    if candidates:
        return candidates[0]
    raise ProcessingError(
        "The document could not be converted. The source file may be corrupted or unsupported."
    )


def pdf_to_word(job_id: str, src: Path, out_dir: Path) -> JobResult:
    """Convert a PDF into an editable Word (.docx) document.

    Uses pdf2docx (PyMuPDF-based) instead of LibreOffice: a PDF opens in
    LibreOffice as a Draw document, which the Writer .docx filter cannot
    save, so a dedicated PDF->DOCX engine is required. Imported lazily so the
    heavy opencv/numpy dependency chain stays out of app startup.
    """
    from pdf2docx import Converter

    out = out_dir / f"{stem(src.name)}.docx"
    converter = Converter(str(src))
    try:
        converter.convert(str(out))
    except Exception as exc:  # noqa: BLE001
        raise ProcessingError(
            "The PDF could not be converted to Word. It may be encrypted, "
            "image-only (scanned), or corrupted."
        ) from exc
    finally:
        converter.close()
    return file_result(job_id, "pdf-to-word", out)


def word_to_pdf(job_id: str, src: Path, out_dir: Path) -> JobResult:
    out = _libreoffice_convert(src, out_dir, "pdf:writer_pdf_Export", ".pdf")
    return file_result(job_id, "word-to-pdf", out)


def excel_to_pdf(job_id: str, src: Path, out_dir: Path) -> JobResult:
    out = _libreoffice_convert(src, out_dir, "pdf:calc_pdf_Export", ".pdf")
    return file_result(job_id, "excel-to-pdf", out)


def ppt_to_pdf(job_id: str, src: Path, out_dir: Path) -> JobResult:
    out = _libreoffice_convert(src, out_dir, "pdf:impress_pdf_Export", ".pdf")
    return file_result(job_id, "ppt-to-pdf", out)


def pdf_to_excel(job_id: str, src: Path, out_dir: Path) -> JobResult:
    """Extract tables from a PDF into an .xlsx workbook (text fallback)."""
    wb = Workbook()
    wb.remove(wb.active)  # start clean
    sheet_count = 0

    with pdfplumber.open(str(src)) as pdf:
        for page_index, page in enumerate(pdf.pages, start=1):
            tables = page.extract_tables() or []
            for table_index, table in enumerate(tables, start=1):
                ws = wb.create_sheet(title=f"P{page_index}T{table_index}"[:31])
                for row in table:
                    ws.append([("" if cell is None else str(cell)) for cell in row])
                sheet_count += 1

        if sheet_count == 0:
            ws = wb.create_sheet(title="Text")
            for page in pdf.pages:
                for line in (page.extract_text() or "").splitlines():
                    ws.append([line])

    if not wb.sheetnames:
        wb.create_sheet(title="Empty")

    out = out_dir / f"{stem(src.name)}.xlsx"
    wb.save(str(out))
    return file_result(job_id, "pdf-to-excel", out, meta={"tables": sheet_count})
